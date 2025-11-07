#!/usr/bin/env python3
"""
Extract unique Australian cities from the AusNails.xlsx file
and prepare them for insertion into the database
"""

import openpyxl
import json
import re
from collections import defaultdict

def slugify(text):
    """Convert text to URL-friendly slug"""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '-', text)
    return text

def extract_cities():
    print("Loading Australian nail salon data...")
    workbook = openpyxl.load_workbook('AusNails.xlsx', read_only=True)
    
    # Try to find the "Salon" sheet
    sheet_name = None
    for name in workbook.sheetnames:
        if 'salon' in name.lower():
            sheet_name = name
            break
    
    if not sheet_name:
        print(f"Available sheets: {workbook.sheetnames}")
        sheet_name = workbook.sheetnames[0]
    
    print(f"Using sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    
    # Find City and State columns
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    
    city_col = None
    state_col = None
    
    for idx, header in enumerate(header_row):
        if header and 'city' in str(header).lower():
            city_col = idx
        if header and 'state' in str(header).lower():
            state_col = idx
    
    print(f"City column: {city_col}, State column: {state_col}")
    
    if city_col is None or state_col is None:
        print("ERROR: Could not find City or State columns")
        print(f"Headers: {header_row[:20]}")
        return
    
    # Extract unique city-state combinations
    cities_by_state = defaultdict(set)
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        city = row[city_col]
        state = row[state_col]
        
        if city and state:
            city = str(city).strip()
            state = str(state).strip()
            
            if city and state:
                cities_by_state[state].add(city)
    
    workbook.close()
    
    # Organize by state
    australian_states = {
        'NSW': 'New South Wales',
        'VIC': 'Victoria',
        'QLD': 'Queensland',
        'WA': 'Western Australia',
        'SA': 'South Australia',
        'TAS': 'Tasmania',
        'ACT': 'Australian Capital Territory',
        'NT': 'Northern Territory'
    }
    
    print("\n" + "="*60)
    print("AUSTRALIAN CITIES BY STATE")
    print("="*60)
    
    all_cities = []
    city_id = 1
    
    for state_code, state_name in australian_states.items():
        cities = sorted(cities_by_state.get(state_code, []))
        
        if cities:
            print(f"\n{state_name} ({state_code}): {len(cities)} cities")
            print("-" * 60)
            
            for city in cities:
                city_data = {
                    'id': city_id,
                    'name': city,
                    'state': state_code,
                    'state_full': state_name,
                    'slug': slugify(city),
                    'country': 'Australia'
                }
                all_cities.append(city_data)
                print(f"  {city_id}. {city} (/{state_code.lower()}/{slugify(city)})")
                city_id += 1
    
    # Save to JSON
    with open('australian_cities.json', 'w') as f:
        json.dump(all_cities, f, indent=2)
    
    print(f"\n{'='*60}")
    print(f"Total cities extracted: {len(all_cities)}")
    print(f"Saved to: australian_cities.json")
    print("="*60)
    
    # Generate SQL for insertion
    print("\n\nGenerating SQL for Supabase...")
    
    with open('insert_australian_cities.sql', 'w') as f:
        f.write("-- Insert Australian cities into the database\n")
        f.write("-- Run this in Supabase SQL Editor\n\n")
        
        # First, ensure we have the Australian states
        f.write("-- Insert Australian states (if not already present)\n")
        for state_code, state_name in australian_states.items():
            f.write(f"INSERT INTO states (code, name, country) VALUES ('{state_code}', '{state_name}', 'Australia') ON CONFLICT (code) DO NOTHING;\n")
        
        f.write("\n-- Insert cities\n")
        for city in all_cities:
            city_name_escaped = city['name'].replace("'", "''")
            f.write(f"INSERT INTO cities (name, state_id, slug, country) SELECT '{city_name_escaped}', id, '{city['slug']}', 'Australia' FROM states WHERE code = '{city['state']}' ON CONFLICT (slug) DO NOTHING;\n")
    
    print("SQL file created: insert_australian_cities.sql")
    
    return all_cities

if __name__ == '__main__':
    cities = extract_cities()
