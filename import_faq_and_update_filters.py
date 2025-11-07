#!/usr/bin/env python3
"""
Import FAQ data and update filter names
"""

import os
import openpyxl
from supabase import create_client, Client

# Initialize Supabase
supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
supabase_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
supabase: Client = create_client(supabase_url, supabase_key)

def import_faq_data():
    """Import About and FAQ text fields from Excel"""
    print("="*80)
    print("IMPORTING FAQ DATA FROM EXCEL")
    print("="*80)
    
    wb = openpyxl.load_workbook('Nail_Salons_Aus_250_updated.xlsx', data_only=True)
    sheet = wb.active
    
    # Get headers
    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col_idx).value
        if header:
            headers[header] = col_idx
    
    print(f"\nProcessing {sheet.max_row - 1} salons...")
    
    updated = 0
    skipped = 0
    
    for row_num in range(2, min(sheet.max_row + 1, 252)):
        # Get salon name
        salon_name = sheet.cell(row=row_num, column=headers.get('name', 2)).value
        if not salon_name:
            skipped += 1
            continue
        
        salon_name = str(salon_name).strip()
        
        # Build FAQ update data
        faq_data = {}
        
        # About
        if 'About' in headers:
            about_val = sheet.cell(row=row_num, column=headers['About']).value
            if about_val:
                faq_data['about'] = str(about_val).strip()
        
        # Review summary
        if 'Reveiw summary' in headers:
            review_val = sheet.cell(row=row_num, column=headers['Reveiw summary']).value
            if review_val:
                faq_data['review_summary'] = str(review_val).strip()
        
        # What are customers saying
        if "What are customer's saying?" in headers:
            customers_val = sheet.cell(row=row_num, column=headers["What are customer's saying?"]).value
            if customers_val:
                faq_data['customers_saying'] = str(customers_val).strip()
        
        # Health and wellbeing
        if 'How do they care for your health and wellbeing?' in headers:
            health_val = sheet.cell(row=row_num, column=headers['How do they care for your health and wellbeing?']).value
            if health_val:
                faq_data['health_wellbeing_care'] = str(health_val).strip()
        
        # Update if we have any FAQ data
        if faq_data:
            try:
                result = supabase.table('salons').update(faq_data).eq('name', salon_name).execute()
                if result.data:
                    updated += 1
                    if updated % 50 == 0:
                        print(f"  ✓ Imported FAQ data for {updated} salons...")
            except Exception as e:
                print(f"  ✗ Error updating {salon_name}: {str(e)[:80]}")
        else:
            skipped += 1
    
    print(f"\n✓ FAQ Import Complete:")
    print(f"  - Updated: {updated} salons")
    print(f"  - Skipped: {skipped} salons")
    print()

def main():
    import_faq_data()
    print("="*80)
    print("✓ ALL OPERATIONS COMPLETED")
    print("="*80)

if __name__ == "__main__":
    main()

