#!/usr/bin/env python3
"""
Import updated salon data from Excel file with all new filter columns and FAQ fields.
Handles column additions and data population in one comprehensive script.
"""

import os
import openpyxl
from supabase import create_client, Client
from typing import Dict, Any, Optional

# Initialize Supabase client
supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
supabase_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

if not supabase_url or not supabase_key:
    raise ValueError("Missing Supabase environment variables")

supabase: Client = create_client(supabase_url, supabase_key)

# Complete column mapping from Excel to database
COLUMN_MAPPING = {
    # Services (AV-BO)
    'Manicure': 'manicure',
    'Gel Manicure': 'gel_manicure',
    'Gel Extensions': 'gel_extensions',
    'Acrylic Nails': 'acrylic_nails',
    'Pedicure': 'pedicure',
    'Gel Pedicure': 'gel_pedicure',
    'SNS Dip Powder': 'sns_dip_powder',
    'Builders Gel / BIAB': 'builders_gel_biab',
    'Nail Art': 'nail_art',
    'Massage': 'massage',
    'Facials': 'facials',
    'Lash Exensions': 'lash_extensions',  # Note: typo in Excel
    'Lash Lift and Tint': 'lash_lift_tint',
    'Brows': 'brows',
    'Waxing': 'waxing',
    'Injectables': 'injectables',
    'Tanning': 'tanning',
    'Cosmetic Tatoo': 'cosmetic_tattoo',  # Note: typo in Excel
    'Haircuts': 'haircuts',
    'Spa Hand and Foot Treatment': 'spa_hand_foot_treatment',
    
    # Price (BP)
    'Price ($-$$$)': 'price_range',
    
    # Languages (BQ-BU)
    'English': 'english',
    'Spanish': 'spanish',
    'Vietnamese': 'vietnamese',
    'Chinese': 'chinese',
    'Korean': 'korean',
    
    # Specialties (BV-CA)
    'Qualified technicians': 'qualified_technicians',
    'Experienced Team': 'experienced_team',
    'Quick Service': 'quick_service',
    'Award winning staff': 'award_winning_staff',
    'Master Nail Artist': 'master_nail_artist',
    'Bridal Nails': 'bridal_nails',
    
    # Appointment Types (CB-CE)
    'Appointment Required': 'appointment_required',
    'Walk-ins Welcome': 'walk_ins_welcome',
    'Group Bookings': 'group_bookings',
    'Mobile Nails': 'mobile_nails',
    
    # Amenities/Space (CF-CS)
    'Child Friendly': 'child_friendly',
    'Adult Only': 'adult_only',
    'Pet Friendly': 'pet_friendly',
    'LGBTQI+ Friendly': 'lgbtqi_friendly',
    'Wheel Chair Accessable': 'wheelchair_accessible',  # Note: typo in Excel
    'Complimentary drink': 'complimentary_drink',
    'Heated Massage Chairs': 'heated_massage_chairs',
    'Foot Spas': 'foot_spas',
    'Free Wi-fi': 'free_wifi',
    'Parking': 'parking',
    'Autoclave sterlisation': 'autoclave_sterilisation',  # Note: typo in Excel
    'LED Curing': 'led_curing',
    'Clean & Ethical Products': 'clean_ethical_products',
    'Vegan Polish': 'vegan_polish',
    
    # FAQ Fields (CT-CX)
    'Reveiw summary': 'review_summary',  # Note: typo in Excel
    'Description': 'description',
    'About': 'about',
    "What are customer's saying?": 'customers_saying',
    'How do they care for your health and wellbeing?': 'health_wellbeing_care',
}

def add_missing_columns():
    """Add any missing columns to the salons table."""
    print("Step 1: Adding missing columns to database...")
    
    # Read SQL migration file
    with open('add_new_columns.sql', 'r') as f:
        sql_commands = f.read()
    
    # Execute each ALTER TABLE statement
    for command in sql_commands.split(';'):
        command = command.strip()
        if command and not command.startswith('--'):
            try:
                result = supabase.rpc('exec_sql', {'query': command}).execute()
                print(f"  ✓ Executed: {command[:60]}...")
            except Exception as e:
                print(f"  ⚠ Warning (may already exist): {str(e)[:80]}")
    
    print("  ✓ Column additions complete\n")

def load_excel_data(filename: str):
    """Load data from Excel file."""
    print(f"Step 2: Loading Excel file: {filename}")
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb.active
    
    # Get headers
    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col_idx).value
        if header:
            headers[header] = col_idx
    
    print(f"  ✓ Found {len(headers)} columns")
    print(f"  ✓ Total rows: {sheet.max_row}\n")
    
    return sheet, headers

def process_salon_row(sheet, row_num: int, headers: Dict[str, int]) -> Optional[Dict[str, Any]]:
    """Process a single salon row and return update data."""
    row = sheet[row_num]
    
    # Get salon name (column B)
    salon_name_col = headers.get('name', 2)
    salon_name = sheet.cell(row=row_num, column=salon_name_col).value
    
    if not salon_name or str(salon_name).strip() == '':
        return None
    
    salon_name = str(salon_name).strip()
    
    # Build update data dictionary
    update_data = {}
    
    # Process all mapped columns
    for excel_col_name, db_col_name in COLUMN_MAPPING.items():
        if excel_col_name not in headers:
            continue
        
        col_idx = headers[excel_col_name]
        cell_value = sheet.cell(row=row_num, column=col_idx).value
        
        # Handle boolean columns (Yes/No)
        if db_col_name != 'price_range' and db_col_name not in ['description', 'review_summary', 'about', 'customers_saying', 'health_wellbeing_care']:
            if isinstance(cell_value, str):
                update_data[db_col_name] = cell_value.strip().lower() == 'yes'
            else:
                update_data[db_col_name] = False
        else:
            # Handle text columns (price_range and FAQ fields)
            if cell_value:
                update_data[db_col_name] = str(cell_value).strip()
    
    return {'name': salon_name, 'data': update_data}

def import_all_data(sheet, headers: Dict[str, int]):
    """Import all salon data from Excel."""
    print("Step 3: Importing salon data...")
    
    total_salons = sheet.max_row - 1  # Minus header row
    updated_count = 0
    skipped_count = 0
    error_count = 0
    
    for row_num in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
        try:
            salon_info = process_salon_row(sheet, row_num, headers)
            
            if not salon_info:
                skipped_count += 1
                continue
            
            salon_name = salon_info['name']
            update_data = salon_info['data']
            
            # Update salon in database
            result = supabase.table('salons').update(update_data).eq('name', salon_name).execute()
            
            if result.data:
                updated_count += 1
                if updated_count % 50 == 0:
                    print(f"  ✓ Processed {updated_count}/{total_salons} salons...")
            else:
                skipped_count += 1
                
        except Exception as e:
            error_count += 1
            if error_count <= 5:  # Only print first 5 errors
                print(f"  ✗ Error processing row {row_num}: {str(e)[:100]}")
    
    print(f"\n  ✓ Import complete!")
    print(f"    - Updated: {updated_count} salons")
    print(f"    - Skipped: {skipped_count} salons")
    print(f"    - Errors: {error_count} salons\n")

def main():
    """Main execution function."""
    print("="*80)
    print("IMPORTING UPDATED SALON DATA WITH NEW FILTERS AND FAQ FIELDS")
    print("="*80)
    print()
    
    try:
        # Step 1: Add missing columns
        add_missing_columns()
        
        # Step 2: Load Excel file
        sheet, headers = load_excel_data('Nail_Salons_Aus_250_updated.xlsx')
        
        # Step 3: Import all data
        import_all_data(sheet, headers)
        
        print("="*80)
        print("✓ ALL OPERATIONS COMPLETED SUCCESSFULLY")
        print("="*80)
        
    except Exception as e:
        print(f"\n✗ FATAL ERROR: {str(e)}")
        raise

if __name__ == "__main__":
    main()

