#!/usr/bin/env python3
"""
Populate service/amenity data from Excel file to Supabase database.
This script reads the Nail_Salons_Aus_250.xlsx file and updates the
boolean service columns in the salons table.
"""

import openpyxl
import os
from supabase import create_client, Client
from dotenv import load_dotenv
import time

# Load environment variables
load_dotenv('.env.local')

# Initialize Supabase client
supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

if not supabase_url or not supabase_key:
    print("❌ Error: Missing Supabase credentials in .env.local")
    exit(1)

supabase: Client = create_client(supabase_url, supabase_key)

print("✅ Connected to Supabase")

# Load Excel file
print("\n📂 Loading Excel file...")
wb = openpyxl.load_workbook('Nail_Salons_Aus_250.xlsx')
sheet = wb.active
headers = [cell.value for cell in sheet[1]]

print(f"✅ Loaded {sheet.max_row - 1} rows from Excel")

# Map Excel column names to database column names
# Format: 'Excel Column Name': 'database_column_name'
column_mapping = {
    # Services (columns AV-BK)
    'c': 'manicure',  # Column AV is just "c" which means Manicure
    'Gel Manicure': 'gel_manicure',
    'Gel X': 'gel_x',
    'Gel Extensions': 'gel_extensions',
    'Acrylic Nails': 'acrylic_nails',
    'Pedicure': 'pedicure',
    'Gel Pedicure': 'gel_pedicure',
    'Dip Powder': 'dip_powder',
    'Builders Gel': 'builders_gel',
    'Nail Art': 'nail_art',
    'Massage': 'massage',
    'Facials': 'facials',
    'Eyelashes': 'eyelashes',
    'Brows': 'brows',
    'Waxing': 'waxing',
    'Hair cuts': 'hair_cuts',
    'Hand and Foot Treatment': 'hand_foot_treatment',
    
    # Languages (columns BN-BS)
    'Basic English': 'basic_english',
    'Fluent English': 'fluent_english',
    'Spanish': 'spanish',
    'Vietnamese': 'vietnamese',
    'Chinese': 'chinese',
    'Korean': 'korean',
    
    # Specialties (columns BT-BY)
    'Qualified technicians': 'certified_technicians',
    'Experienced Team': 'experienced_staff',
    'Quick Service': 'quick_service',
    'Award winning staff': 'award_winning_staff',
    'Master Nail Artist': 'master_artist',
    'Bridal Nails': 'bridal_nails',
    
    # Booking (columns BZ-CC)
    'Appointment Required': 'appointment_only',
    'Walk-ins Welcome': 'accepts_walk_ins',
    'Group Bookings': 'group_bookings',
    'Mobile Nails': 'mobile_nails',
    
    # Amenities (columns CD-CV)
    'Kid friendly': 'kid_friendly',
    'Child play area': 'child_play_area',
    'Adult only': 'adult_only',
    'Pet friendly': 'pet_friendly',
    'LGBTQI+ friendly': 'lgbtqi_friendly',
    'Wheel chair accessable': 'wheelchair_accessible',
    'Female owned salon': 'female_owned',
    'Minority owned salon': 'minority_owned',
    'Complimentary drink': 'complimentary_drink',
    'Heated Massage Chairs': 'heated_massage_chairs',
    'Foot Spas': 'foot_spas',
    'Free Wi-fi': 'free_wifi',
    'Parking': 'parking',
    'Autoclave sterlisation': 'autoclave_sterilisation',
    'LED curing': 'led_curing',
    'Non-toxic treatments': 'non_toxic_treatments',
    'Eco-friendly products': 'eco_friendly_products',
    'Cruelty free products': 'cruelty_free_products',
    'Vegan polish': 'vegan_polish',
}

# Find column indices for each mapped column
excel_col_indices = {}
for idx, header in enumerate(headers, 1):
    if header in column_mapping:
        excel_col_indices[header] = idx
        print(f"  ✓ Found '{header}' at column {idx}")

print(f"\n✅ Mapped {len(excel_col_indices)} columns")

# Process each row
success_count = 0
error_count = 0
skip_count = 0
batch_size = 10
batch_updates = []

print(f"\n🔄 Processing {sheet.max_row - 1} salons...")

for row_num in range(2, sheet.max_row + 1):
    row = sheet[row_num]
    salon_name = row[1].value  # Column B is name
    
    if not salon_name:
        skip_count += 1
        continue
    
    # Build update data for this salon
    update_data = {}
    
    for excel_col_name, col_idx in excel_col_indices.items():
        cell_value = row[col_idx - 1].value  # -1 because row is 0-indexed
        db_col_name = column_mapping[excel_col_name]
        
        # Convert 'Yes' to True, anything else to False
        if isinstance(cell_value, str) and cell_value.strip().lower() == 'yes':
            update_data[db_col_name] = True
        else:
            update_data[db_col_name] = False
    
    # Add to batch
    batch_updates.append({
        'name': salon_name,
        'update_data': update_data
    })
    
    # Process batch when it reaches batch_size
    if len(batch_updates) >= batch_size:
        try:
            for item in batch_updates:
                result = supabase.table('salons').update(item['update_data']).eq('name', item['name']).execute()
                if result.data:
                    success_count += 1
                else:
                    error_count += 1
                    print(f"  ⚠️  No match for: {item['name']}")
            
            print(f"  ✓ Processed batch: {success_count} successful, {error_count} errors")
            batch_updates = []
            time.sleep(0.1)  # Small delay to avoid rate limiting
            
        except Exception as e:
            print(f"  ❌ Batch error: {str(e)}")
            error_count += len(batch_updates)
            batch_updates = []

# Process remaining batch
if batch_updates:
    try:
        for item in batch_updates:
            result = supabase.table('salons').update(item['update_data']).eq('name', item['name']).execute()
            if result.data:
                success_count += 1
            else:
                error_count += 1
                print(f"  ⚠️  No match for: {item['name']}")
        
        print(f"  ✓ Processed final batch: {success_count} successful, {error_count} errors")
        
    except Exception as e:
        print(f"  ❌ Final batch error: {str(e)}")
        error_count += len(batch_updates)

print("\n" + "="*60)
print("📊 IMPORT SUMMARY")
print("="*60)
print(f"✅ Successfully updated: {success_count} salons")
print(f"❌ Errors: {error_count}")
print(f"⏭️  Skipped (no place_id): {skip_count}")
print(f"📁 Total rows processed: {sheet.max_row - 1}")
print("="*60)

# Verify the import
print("\n🔍 Verifying import...")
try:
    # Check how many salons now have manicure=true
    result = supabase.table('salons').select('id, manicure, gel_manicure, pedicure').eq('manicure', True).limit(5).execute()
    
    if result.data:
        print(f"✅ Found {len(result.data)} salons with manicure=True (showing first 5)")
        for salon in result.data:
            print(f"  - Salon ID {salon['id']}: manicure={salon['manicure']}, gel_manicure={salon['gel_manicure']}, pedicure={salon['pedicure']}")
    else:
        print("⚠️  No salons found with manicure=True - checking if update worked...")
        
    # Check total count
    total = supabase.table('salons').select('id', count='exact').execute()
    print(f"\n📊 Total salons in database: {total.count}")
    
except Exception as e:
    print(f"❌ Verification error: {str(e)}")

print("\n✅ Script completed!")
