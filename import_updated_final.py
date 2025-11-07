#!/usr/bin/env python3
"""
Final import script with proper price range handling
"""

import os
import openpyxl
from supabase import create_client, Client

# Initialize Supabase client
supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
supabase_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

if not supabase_url or not supabase_key:
    raise ValueError("Missing Supabase environment variables")

supabase: Client = create_client(supabase_url, supabase_key)

# Complete column mapping (Excel name -> DB name)
COLUMN_MAPPING = {
    # Services
    'Manicure': 'manicure',
    'Gel Manicure': 'gel_manicure',
    'Gel Extensions': 'gel_extensions',
    'Acrylic Nails': 'acrylic_nails',
    'Pedicure': 'pedicure',
    'Gel Pedicure': 'gel_pedicure',
    'SNS Dip Powder': 'sns_dip_powder',
    'Builders Gel / BIAB': 'builders_gel_biab',
    'Nail Art': 'nail_art',
    'Massage': 'massage',
    'Facials': 'facials',
    'Lash Exensions': 'lash_extensions',
    'Lash Lift and Tint': 'lash_lift_tint',
    'Brows': 'brows',
    'Waxing': 'waxing',
    'Injectables': 'injectables',
    'Tanning': 'tanning',
    'Cosmetic Tatoo': 'cosmetic_tattoo',
    'Haircuts': 'haircuts',
    'Spa Hand and Foot Treatment': 'spa_hand_foot_treatment',
    # Languages
    'English': 'english',
    'Spanish': 'spanish',
    'Vietnamese': 'vietnamese',
    'Chinese': 'chinese',
    'Korean': 'korean',
    # Specialties
    'Qualified technicians': 'qualified_technicians',
    'Experienced Team': 'experienced_team',
    'Quick Service': 'quick_service',
    'Award winning staff': 'award_winning_staff',
    'Master Nail Artist': 'master_nail_artist',
    'Bridal Nails': 'bridal_nails',
    # Appointment Types
    'Appointment Required': 'appointment_required',
    'Walk-ins Welcome': 'walk_ins_welcome',
    'Group Bookings': 'group_bookings',
    'Mobile Nails': 'mobile_nails',
    # Amenities
    'Child Friendly': 'child_friendly',
    'Adult Only': 'adult_only',
    'Pet Friendly': 'pet_friendly',
    'LGBTQI+ Friendly': 'lgbtqi_friendly',
    'Wheel Chair Accessable': 'wheelchair_accessible',
    'Complimentary drink': 'complimentary_drink',
    'Heated Massage Chairs': 'heated_massage_chairs',
    'Foot Spas': 'foot_spas',
    'Free Wi-fi': 'free_wifi',
    'Parking': 'parking',
    'Autoclave sterlisation': 'autoclave_sterilisation',
    'LED Curing': 'led_curing',
    'Clean & Ethical Products': 'clean_ethical_products',
    'Vegan Polish': 'vegan_polish',
}

def normalize_price_range(value):
    """Convert price range to valid format"""
    if not value:
        return None
    
    value_str = str(value).strip().lower()
    
    # Map various formats to valid values
    if value_str in ['$', 'budget', 'low', 'cheap']:
        return 'budget'
    elif value_str in ['$$', 'mid', 'medium', 'moderate', 'mid-range']:
        return 'mid-range'
    elif value_str in ['$$$', 'high', 'expensive', 'premium', 'luxury']:
        return 'premium'
    
    return None

def load_excel_data(filename: str):
    print(f"Loading Excel file: {filename}")
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb.active
    
    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col_idx).value
        if header:
            headers[header] = col_idx
    
    print(f"  ✓ Found {len(headers)} columns")
    print(f"  ✓ Total rows: {sheet.max_row}")
    return sheet, headers

def import_all_data(sheet, headers):
    print("\nImporting salon data...")
    
    updated_count = 0
    skipped_count = 0
    error_count = 0
    
    for row_num in range(2, min(sheet.max_row + 1, 252)):  # Only first 250 data rows
        try:
            # Get salon name
            salon_name_col = headers.get('name', 2)
            salon_name = sheet.cell(row=row_num, column=salon_name_col).value
            
            if not salon_name or str(salon_name).strip() == '':
                skipped_count += 1
                continue
            
            salon_name = str(salon_name).strip()
            
            # Build update data
            update_data = {}
            
            for excel_col_name, db_col_name in COLUMN_MAPPING.items():
                if excel_col_name not in headers:
                    continue
                
                col_idx = headers[excel_col_name]
                cell_value = sheet.cell(row=row_num, column=col_idx).value
                
                # Boolean columns
                if isinstance(cell_value, str):
                    update_data[db_col_name] = cell_value.strip().lower() == 'yes'
                else:
                    update_data[db_col_name] = False
            
            # Handle price range separately
            if 'Price ($-$$$)' in headers:
                price_col = headers['Price ($-$$$)']
                price_value = sheet.cell(row=row_num, column=price_col).value
                normalized_price = normalize_price_range(price_value)
                if normalized_price:
                    update_data['price_range'] = normalized_price
            
            # Update salon
            result = supabase.table('salons').update(update_data).eq('name', salon_name).execute()
            
            if result.data:
                updated_count += 1
                if updated_count % 50 == 0:
                    print(f"  ✓ Processed {updated_count} salons...")
            else:
                skipped_count += 1
                
        except Exception as e:
            error_count += 1
            if error_count <= 3:
                error_msg = str(e)
                if len(error_msg) > 100:
                    error_msg = error_msg[:100] + "..."
                print(f"  ✗ Error on row {row_num} ({salon_name if 'salon_name' in locals() else 'unknown'}): {error_msg}")
    
    print(f"\n  ✓ Import complete!")
    print(f"    - Updated: {updated_count} salons")
    print(f"    - Skipped: {skipped_count} salons")
    print(f"    - Errors: {error_count} salons")

def main():
    print("="*80)
    print("FINAL DATA IMPORT - Services, Languages, Specialties, Amenities & Price Range")
    print("="*80)
    print()
    
    try:
        sheet, headers = load_excel_data('Nail_Salons_Aus_250_updated.xlsx')
        import_all_data(sheet, headers)
        
        print("\n" + "="*80)
        print("✓ IMPORT COMPLETED SUCCESSFULLY")
        print("="*80)
        
    except Exception as e:
        print(f"\n✗ FATAL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

if __name__ == "__main__":
    main()

